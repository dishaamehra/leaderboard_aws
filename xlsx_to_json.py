"""
Run this after updating IMS_Hackathon_Leaderboard.xlsx each week.
It reads the Leaderboard sheet's calculated values and writes data.json,
which the website reads directly. No personal info (emails/phones) is
ever exported -- only team names, tracks, and points.

Usage:
    python3 xlsx_to_json.py
"""
import json
import datetime
from openpyxl import load_workbook

XLSX_PATH = "IMS_Hackathon_Leaderboard.xlsx"
OUT_PATH = "site/data.json"
EVENT_NAME = "BuildQuest Leaderboard"
NUM_WEEKS = 6

wb = load_workbook(XLSX_PATH, data_only=True)
ws = wb["Leaderboard"]
rows = list(ws.iter_rows(min_row=2, values_only=True))

teams = []
for r in rows:
    rank, team_id, team_name, track = r[0], r[1], r[2], r[3]
    weekly = [float(v or 0) for v in r[4:4 + NUM_WEEKS]]
    total = float(r[4 + NUM_WEEKS] or 0)
    teams.append({
        "rank": int(rank),
        "team_name": team_name,
        "track": track,
        "weekly": weekly,
        "total": total,
    })

teams.sort(key=lambda t: t["rank"])

# auto-detect the latest week that has any activity logged
current_week = 0
for wk in range(NUM_WEEKS, 0, -1):
    if any(t["weekly"][wk - 1] > 0 for t in teams):
        current_week = wk
        break

data = {
    "event_name": EVENT_NAME,
    "num_weeks": NUM_WEEKS,
    "current_week": current_week,
    "last_updated": datetime.date.today().isoformat(),
    "teams": teams,
}

import os
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
with open(OUT_PATH, "w") as f:
    json.dump(data, f, indent=2)

print(f"Wrote {OUT_PATH} -- {len(teams)} teams, current_week={current_week}")
